#!/usr/bin/env python3
"""Independent QA for owner-authorized Step18/19 corrections discovered by Step20."""
from __future__ import annotations

import hashlib
import json
import zipfile
from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "STEP_19_CLIENT_WORKBOOK_CORRECTED.xlsx"
DOCX = ROOT / "STEP_19_CLIENT_REPORT_CORRECTED.docx"
PDF = ROOT / "STEP_19_CLIENT_REPORT_CORRECTED.pdf"
MANIFEST = ROOT / "STEP_19_PHYSICAL_ARTIFACT_MANIFEST_2026-09-03.json"


def sha(path: Path):
    h = hashlib.sha256(); h.update(path.read_bytes()); return h.hexdigest()


# Workbook identity/content.
wb = load_workbook(XLSX, read_only=True, data_only=False)
expected = ["README", "02_Page_Model", "03_Semantic_Core", "04_Search_vs_AI", "05_Page_Actions", "06_Source_Obs", "07_Priority_Plan", "Execution_Calibration", "Measurement"]
assert wb.sheetnames == expected
assert "TEST/DEMO CASE" in str(wb["README"]["A1"].value)
assert "mock commercial rehearsal" in str(wb["README"]["A2"].value)
assert wb["03_Semantic_Core"].max_row - 1 == 2332
assert wb["Execution_Calibration"].max_row - 1 == 112
assert wb["Measurement"].max_row - 1 == 7

hdr = [c.value for c in wb["05_Page_Actions"][1]]
rows = [dict(zip(hdr, [c.value for c in r])) for r in wb["05_Page_Actions"].iter_rows(min_row=2)]
a12 = next(r for r in rows if r["action_id"] == "S18-A012")
a27 = next(r for r in rows if r["action_id"] == "S18-A027")
assert "Сохранить уже существующую price/price-estimation guidance" in a12["client_action"]
assert "Не добавлять повторное базовое определение" in a27["client_action"]
assert "Не дублировать уже существующие price factors" in a12["do_not_do_boundary"]
assert "Не создавать отдельный дублирующий definition block" in a27["do_not_do_boundary"]

ph = [c.value for c in wb["07_Priority_Plan"][1]]
pr = [dict(zip(ph, [c.value for c in r])) for r in wb["07_Priority_Plan"].iter_rows(min_row=2)]
p12 = next(r for r in pr if r["action_id"] == "S18-A012")
p27 = next(r for r in pr if r["action_id"] == "S18-A027")
assert "price/price-estimation guidance is already present" in p12["client_readable_reason"]
assert "basic French-window definition is already present" in p27["client_readable_reason"]

ch = [c.value for c in wb["Execution_Calibration"][1]]
cr = [dict(zip(ch, [c.value for c in r])) for r in wb["Execution_Calibration"].iter_rows(min_row=2)]
c12 = next(r for r in cr if r["source_action_id"] == "S18-A012")
c27 = next(r for r in cr if r["source_action_id"] == "S18-A027")
assert c12["what_to_do"].startswith("Retain the current door price/price-estimation guidance")
assert c27["what_to_do"].startswith("Do not add a duplicate basic French-window definition")
assert c27["depends_on_action_ids"] == "S18-A009"
wb.close()

# DOCX visible identity/content + real core-property hygiene.
doc = Document(DOCX)
text = "\n".join(p.text for p in doc.paragraphs)
assert "TEST/DEMO CASE" in text and "mock commercial rehearsal" in text
assert "Сохранить уже существующую price/price-estimation guidance" in text
assert (doc.core_properties.author or "") != "python-docx"
assert (doc.core_properties.author or "") == ""
assert doc.core_properties.created is not None and doc.core_properties.created.year == 2026
assert doc.core_properties.modified is not None and doc.core_properties.modified.year == 2026
with zipfile.ZipFile(DOCX) as zf:
    document_xml = zf.read("word/document.xml").decode("utf-8")
    core_xml = zf.read("docProps/core.xml").decode("utf-8")
assert "w:cantSplit" in document_xml and "w:tblHeader" in document_xml
assert "python-docx" not in core_xml
assert "2013-12-23" not in core_xml
assert "2026-09-03" in core_xml

# PDF visible demo identity.
reader = PdfReader(PDF)
assert len(reader.pages) >= 3
pdf_text = "\n".join((pg.extract_text() or "") for pg in reader.pages)
assert "TEST/DEMO CASE" in pdf_text and "Mock commercial rehearsal" in pdf_text

# Manifest and exact byte identities.
manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
assert manifest["distribution_identity"]["test_demo_case"] is True
assert manifest["distribution_identity"]["mock_commercial_rehearsal"] is True
assert manifest["distribution_identity"]["actual_paid_client_engagement"] is False
for f in (XLSX, DOCX, PDF):
    assert manifest["artifacts"][f.name]["bytes"] == f.stat().st_size
    assert manifest["artifacts"][f.name]["sha256"] == sha(f)

qa = {
    "date": "2026-09-03",
    "status": "PASS_STEP18_STEP19_CORRECTION_FROM_STEP20_DEFECTS",
    "resolved": {
        "D20-001": "TEST_DEMO_VISIBLE_IN_XLSX_DOCX_PDF_AND_MANIFEST",
        "D20-002": "A012_NARROWED_TO_INSTALLATION_SCOPE_PROCESS_EXISTING_PRICE_GUIDANCE_RETAINED",
        "D20-003": "A027_NO_DUPLICATE_DEFINITION_RESIDUAL_DISTINCTION_FOLDED_WITH_A009",
        "D20-004": "DOCX_CORE_METADATA_CLEANED_TO_2026_NO_PYTHON_DOCX_CREATOR",
    },
    "accounting": {
        "semantic_rows": 2332,
        "actions": 34,
        "execution_packages": 112,
        "hold_packages": 20,
        "measurement_classes": 7,
    },
    "provider_calls": 0,
    "new_paid_cost_rub": 0.0,
    "physical_artifacts": manifest["artifacts"],
    "step20_rerun_required": True,
    "step21_allowed": False,
}
(ROOT / "STEP_18_STEP19_STEP20_DEFECT_CORRECTION_QA_2026-09-03.json").write_text(json.dumps(qa, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
print("STEP18_STEP19_STEP20_DEFECT_CORRECTION_QA_PASS")
